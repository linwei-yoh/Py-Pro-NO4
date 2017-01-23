#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
import re
import pickle

tar_excel_path = '../xlsx/target.xlsx'
output_path = '../xlsx/output.xlsx'
pick_path = '../pickle/save.txt'


# 1.逐个读取各个专利 获取专利名称(唯一） 和 专利所有人
# 2.查询当前专利所有人是否已经在所有人字典中存在 否，则创建并添加 是,则添加专利名称到其专利集合中
# 3.遍历完成，获得所有人字典 记载了全部所有人，及其所拥有的专利
# 4.逐个读取各个专利 获取专利名称(唯一） 和 引用专利(转化为唯一专利）
# 5.查询引用专利是否在专利字典中存在 否，则创建并添加 是,则添加专利名称到引用专利的被引用专利集合中
# 6.遍历完成，获得专利被引用字典 记载了全部被引用专利，及其所引用该专利的专利集合
# 7.逐个遍历所有人字典，对单个所有人的全部专利的被引用集合作并集处理，获得对此所有人的引用
# 8.获得该并集与有该所有人拥有的专利集合的差集(集合相减)，获得有效的专利引用集合，获得中心度
# 9.增加所有人的集合体 公司，专利对象增加 引用集合属性

class GroupItem(object):
    """用于包含公司的集团对象"""

    def __init__(self, name):
        self.name = name
        self.company_set = set()

    def add_company(self, name):
        self.company_set.add(name)


class CompanyItem(object):
    """所有人对象 包含所有人名称 和 所有专利集合"""

    def __init__(self, name):
        self.name = name
        self.patent_set = set()
        self.group_name = 'None'

    def add_patent(self, name):
        self.patent_set.add(name)


class PatentItem(object):
    """专利对象 包含专利名称 和 引用专利集合"""

    def __init__(self, name):
        self.name = name
        self.cited_set = set()
        self.cite_set = set()
        self.owner_set = set()

    # 该专利被引用专利集合
    def add_cited_set(self, name):
        self.cited_set.add(name)

    # 该专利引用集合
    def add_cite_set(self, name):
        self.cite_set.add(name)

    def add_owner(self, company_name):
        self.owner_set.add(company_name)


group_dict = {}  # 集团字典
company_dict = {}  # 所有人字典
patent_dict = {}  # 专利字典
unique_name_dict = {}  # 专利唯一名称字典


def add_one_info(group_name, company_name, patent_name):
    '''建立集团对象 公司对象 专利对象'''
    global group_dict
    global company_dict
    global patent_dict

    if group_name not in group_dict:
        new_group = GroupItem(group_name)
        group_dict[group_name] = new_group
    group_item = group_dict[group_name]

    if company_name not in company_dict:
        new_company = CompanyItem(company_name)
        company_dict[company_name] = new_company
    company_item = company_dict[company_name]

    if patent_name not in patent_dict:
        new_patent = PatentItem(patent_name)
        patent_dict[patent_name] = new_patent

    group_item.add_company(company_name)
    company_item.group_name = group_name
    company_item.add_patent(patent_name)
    patent_dict[patent_name].add_owner(company_name)


def add_patent_and_cite(name_patent, cite_ver):
    """建立专利-引用关系"""
    global patent_dict

    # 关联引用
    if name_patent in patent_dict:
        patent_dict[name_patent].add_cite_set(cite_ver)


def add_patent_and_cited(name_patent, cited_ver):
    """ 建立专利-被引用关系"""
    global patent_dict

    # 关联被引用
    if name_patent in patent_dict:
        patent_dict[name_patent].add_cited_set(cited_ver)


def set_unique_name(name_set, unique_name):
    """建立专利号-唯一专利号字典"""
    global unique_name_dict
    for name_item in name_set:
        if name_item not in unique_name_dict:
            unique_name_dict[name_item] = unique_name


def get_unique_name(names):
    """通过专利号获得对应的唯一专利号"""
    if names in unique_name_dict:
        return unique_name_dict[names]
    else:
        return None


def init_data_from_excel(excel_path):
    """读取excel数据，并建立数据关联用的字典"""
    global group_dict
    global company_dict
    global patent_dict

    print('正在读取excel文件.................')
    wb = load_workbook(excel_path)
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])
    print('读取完成,开始数据处理.............')

    # 获得行数
    num_rows = ws.max_row

    # 获得标签行
    tag_row = [item.value for item in ws[1]]
    tag_row = [item.strip() for item in tag_row]

    # 获得关键数据列号
    patent_index = tag_row.index('GA') + 1  # 专利唯一名称
    company_index = tag_row.index('AE') + 1  # 所有人名称
    ref_index = tag_row.index('CP') + 1  # 引用专利
    subname_index = tag_row.index('PN') + 1  # 专利多个名称

    pattern_com = re.compile(r'\((.+?)\)')
    # 建立 所有人-唯一专利名字典 专利名-唯一专利名字典
    #      集团-所有人
    for i in range(2, num_rows):
        # 公司set
        company_str = str(ws.cell(row=i, column=company_index).value)
        company_set = set([item.strip() for item in company_str.split(';')])
        # 专利唯一名称
        patent_name = str(ws.cell(row=i, column=patent_index).value)
        # 专利子名称set
        subname_str = str(ws.cell(row=i, column=subname_index).value)
        subname_set = set([item.strip() for item in subname_str.split(';')])

        # 建立专利名-唯一专利名字典
        set_unique_name(subname_set, patent_name)

        # 建立集团 所有人 唯一专利名对象 字典 对象关联
        for company_name in company_set:
            group_name = pattern_com.findall(company_name)[0]
            add_one_info(group_name, company_name, patent_name)

        print('对象集合建立处理已经完成%d/%d' % (i, num_rows))

    # 建立专利-引用关系
    print("正在处理专利引用关系")
    pattern = re.compile(r'([a-zA-Z0-9]+?-[A-Za-z]\d?)\s')
    for j in range(2, num_rows):
        # 专利名
        patent_name = str(ws.cell(row=j, column=patent_index).value)
        # 引用专利名集合
        cite_set = set()
        cite_str = str(ws.cell(row=j, column=ref_index).value)
        if cite_str is None or cite_str.strip() == '':
            pass
        else:
            # 获取匹配到的引用集合
            cite_set = set(pattern.findall(cite_str)[1:])

        # 遍历引用集合 建立专利间的引用和被引用关系
        for name in cite_set:
            unique_name = get_unique_name(name)
            # 引用了自身则跳过
            if unique_name == patent_name:
                continue

            # 引用的唯一专利号不存在 则只记录引用关系
            if unique_name is None:
                add_patent_and_cite(patent_name, name)
            else:
                add_patent_and_cite(patent_name, unique_name)
                add_patent_and_cited(unique_name, patent_name)

    print('专利引用关系处理完成')


def pickle_save_dicts(path):
    global group_dict
    global company_dict
    global patent_dict

    print("准备保存关系字典")
    with open(path, 'wb') as save_dicts:
        pickle.dump(group_dict, save_dicts)
        pickle.dump(company_dict, save_dicts)
        pickle.dump(patent_dict, save_dicts)
        pickle.dump(unique_name_dict, save_dicts)
    print("保存关系字典完成")


def pickle_read_dicts(path):
    global group_dict
    global company_dict
    global patent_dict
    global unique_name_dict

    print("准备读取关系字典")
    with open(path, 'rb') as read_dict:
        group_dict = pickle.load(read_dict)
        company_dict = pickle.load(read_dict)
        patent_dict = pickle.load(read_dict)
        unique_name_dict = pickle.load(read_dict)
    print("读取关系字典完成")


def create_sheet_one(wb):
    """建立公司中心度表"""
    ws1 = wb.create_sheet("公司中心度")

    sheet_row_index = 2
    group_name_index = 1
    company_name_index = 2
    degree_name_index = 3

    ws1.cell(row=1, column=group_name_index, value='集团名称')
    ws1.cell(row=1, column=company_name_index, value='公司名称')
    ws1.cell(row=1, column=degree_name_index, value='中心度')

    for group_name, group_item in group_dict.items():
        for company_name in group_item.company_set:
            cited_list = []
            # 获得该公司 所有专利的被引用记录列表
            company_item = company_dict[company_name]
            for patent_name in company_item.patent_set:
                cited_list += list(patent_dict[patent_name].cited_set)
            # 去除自身公司的专利
            valid_list = [ele for ele in cited_list if ele not in company_item.patent_set]

            ws1.cell(row=sheet_row_index, column=group_name_index, value=group_name)
            ws1.cell(row=sheet_row_index, column=company_name_index, value=company_name)
            ws1.cell(row=sheet_row_index, column=degree_name_index, value=len(valid_list))
            sheet_row_index += 1

    print("完成中心度表")


def create_sheet_two(wb):
    """建立专利引用表"""
    ws2 = wb.create_sheet("专利引用表")
    sheet_row_index = 2
    patent_index = 1
    patent_cited_index = 2

    ws2.cell(row=1, column=patent_index, value='唯一专利号')
    ws2.cell(row=1, column=patent_cited_index, value='引用专利号')

    for patent_name, patent_item in patent_dict.items():
        if len(patent_item.cite_set) == 0:
            ws2.cell(row=sheet_row_index, column=patent_index, value=patent_name)
            ws2.cell(row=sheet_row_index, column=patent_cited_index, value='')
            sheet_row_index += 1
        else:
            for item_cite in patent_item.cite_set:
                ws2.cell(row=sheet_row_index, column=patent_index, value=patent_name)
                ws2.cell(row=sheet_row_index, column=patent_cited_index, value=item_cite)
                sheet_row_index += 1

    print("完成专利引用表")


def create_sheet_two_without_unknow(wb):
    ws2_1 = wb.create_sheet("有效专利引用表")
    sheet_row_index = 2
    patent_index = 1
    patent_cited_index = 2

    ws2_1.cell(row=1, column=patent_index, value='唯一专利号')
    ws2_1.cell(row=1, column=patent_cited_index, value='引用专利号')

    for patent_name, patent_item in patent_dict.items():
        if len(patent_item.cite_set) == 0:
            ws2_1.cell(row=sheet_row_index, column=patent_index, value=patent_name)
            ws2_1.cell(row=sheet_row_index, column=patent_cited_index, value='')
            sheet_row_index += 1
        else:
            for item_cite in patent_item.cite_set:
                if item_cite in patent_dict:
                    ws2_1.cell(row=sheet_row_index, column=patent_index, value=patent_name)
                    ws2_1.cell(row=sheet_row_index, column=patent_cited_index, value=item_cite)
                    sheet_row_index += 1

    print("完成有效专利引用表")


def create_sheet_three(wb):
    """公司引用表"""
    ws3 = wb.create_sheet("公司引用表")

    sheet_row_index = 2
    c_name_index = 1
    ref_name_index = 2
    count_index = 3

    ws3.cell(row=1, column=c_name_index, value='公司名')
    ws3.cell(row=1, column=ref_name_index, value='引用公司名')
    ws3.cell(row=1, column=count_index, value='引用次数')

    # 遍历每一个公司
    for company_name, company_item in company_dict.items():
        cite_recode = {}
        tmp_list = []
        # 遍历它的每个专利 获得该专利的引用专利列表
        for patent_name in company_item.patent_set:
            tmp_list += list(patent_dict[patent_name].cite_set)

        # 通过引用专利列表获得引用公司列表
        com_list = []
        for p_name in tmp_list:
            if p_name in patent_dict:
                com_list += list(patent_dict[p_name].owner_set)

        # 获得对其他公司的引用统计字典
        for c_name in com_list:
            if c_name != company_name:
                if c_name not in cite_recode:
                    cite_recode[c_name] = 1
                else:
                    cite_recode[c_name] += 1

        for c_name, count in cite_recode.items():
            ws3.cell(row=sheet_row_index, column=c_name_index, value=company_name)
            ws3.cell(row=sheet_row_index, column=ref_name_index, value=c_name)
            ws3.cell(row=sheet_row_index, column=count_index, value=count)
            sheet_row_index += 1

    print("完成公司引用表")


def create_sheet_four(wb):
    """集团引用表"""
    ws4 = wb.create_sheet("集团引用表")
    sheet_row_index = 2

    g_name_index = 1
    ref_name_index = 2
    count_index = 3

    ws4.cell(row=1, column=g_name_index, value='集团名')
    ws4.cell(row=1, column=ref_name_index, value='引用集团名')
    ws4.cell(row=1, column=count_index, value='引用次数')

    # 遍历集团
    for group_name, group_item in group_dict.items():
        cite_recode = {}
        com_list = []
        # 遍历集团的公司
        for c_name in group_item.company_set:
            tmp_list = []
            # 遍历它的每个专利 获得该专利的引用专利列表
            for patent_name in company_dict[c_name].patent_set:
                tmp_list += list(patent_dict[patent_name].cite_set)

            # 通过引用专利列表获得引用公司列表
            for p_name in tmp_list:
                if p_name in patent_dict:
                    com_list += list(patent_dict[p_name].owner_set)

        # 遍历引用公司列表
        for c_name in com_list:
            # 筛选出不属于本集团的公司 并计数 获得本公司的统计字典
            g_name = company_dict[c_name].group_name
            if g_name != group_name:
                if g_name not in cite_recode:
                    cite_recode[g_name] = 1
                else:
                    cite_recode[g_name] += 1

        for g_name, count in cite_recode.items():
            ws4.cell(row=sheet_row_index, column=g_name_index, value=group_name)
            ws4.cell(row=sheet_row_index, column=ref_name_index, value=g_name)
            ws4.cell(row=sheet_row_index, column=count_index, value=count)
            sheet_row_index += 1
    print("完成集团引用表")


def save_to_excel(path):
    print("开始创建excel表格")
    wb = Workbook()

    # 表1 公司中心度
    create_sheet_one(wb)

    # 表2 专利引用表
    create_sheet_two(wb)
    create_sheet_two_without_unknow(wb)

    # 表3 公司引用表
    create_sheet_three(wb)

    # 表4 集团引用
    create_sheet_four(wb)

    wb.save(path)
    print('excel 创建完成')


if __name__ == '__main__':
    # 只用运行其中一个部分即可
    # 第一部分用于读取与创建数据关联
    # init_data_from_excel(tar_excel_path)
    # pickle_save_dicts(pick_path)
    # 第二部分用于数据统计处理
    pickle_read_dicts(pick_path)
    save_to_excel(output_path)
