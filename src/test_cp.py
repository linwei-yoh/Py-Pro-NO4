#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
import re

target_path = '../xlsx/target.xlsx'
output_path = '../xlsx/output.xlsx'
CP_Index = 18  # 从1开始计算


def test_fun():
    # 准备读的excel
    print('开始读取excel文件')
    wb = load_workbook(target_path)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    print('读取完成')

    # 获得行数
    num_rows = ws.max_row

    # 用于写的sheet
    wb_write = Workbook()
    ws_write = wb_write.active

    row_index = 1
    pattern = re.compile(r'([a-zA-Z0-9]+?-[A-Za-z]{1}\d?)\s')

    # 逐行读数据
    for i in range(2, num_rows):
        CP_val = ws.cell(row=i, column=CP_Index).value
        # CP列存在有效数据
        if CP_val == None or CP_val.strip() == '':
            continue
        else:
            res_list = pattern.findall(CP_val)
            if len(res_list) < 1:
                continue
            else:
                for i in range(1, len(res_list)):
                    ws_write.cell(row=row_index, column=i, value=res_list[i])
                row_index += 1

    wb_write.save(output_path)


if __name__ == '__main__':
    test_fun()
