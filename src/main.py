#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
import re

tar_excel_path = '../xlsx/target.xlsx'
output_path = '../xlsx/output.xlsx'


class GroupItem(object):
    def __init__(self, name):
        self.name = name
        self.company_set = set()

    def add_company(self, name):
        self.company_set.add(name)


# 所有人对象 包含所有人名称 和 所有专利集合
class CompanyItem(object):
    def __init__(self, name):
        self.name = name
        self.patent_set = set()
        self.centre_degrees = 0

    def add_patent(self, name):
        self.patent_set.add(name)


# 专利对象 包含专利名称 和 引用专利集合
class PatentItem(object):
    def __init__(self, name):
        self.name = name
        self.cited_set = set()
        self.cite_set = set()

    # 该专利被引用专利集合
    def add_cited_set(self, name):
        self.cited_set.add(name)

    # 该专利引用集合
    def add_cite_set(self, name):
        if isinstance(name, str):
            self.cite_set.add(name)
        elif isinstance(name, set):
            self.cite_set.update(name)


group_dict = {}  # 集团字典
company_dict = {}  # 所有人字典
patent_dict = {}  # 专利字典
unique_name_dict = {}  # 专利唯一名称字典


# 1.逐个读取各个专利 获取专利名称(唯一） 和 专利所有人
# 2.查询当前专利所有人是否已经在所有人字典中存在 否，则创建并添加 是,则添加专利名称到其专利集合中
# 3.遍历完成，获得所有人字典 记载了全部所有人，及其所拥有的专利
# 4.逐个读取各个专利 获取专利名称(唯一） 和 引用专利(转化为唯一专利）
# 5.查询引用专利是否在专利字典中存在 否，则创建并添加 是,则添加专利名称到引用专利的被引用专利集合中
# 6.遍历完成，获得专利被引用字典 记载了全部被引用专利，及其所引用该专利的专利集合
# 7.逐个遍历所有人字典，对单个所有人的全部专利的被引用集合作并集处理，获得对此所有人的引用
# 8.获得该并集与有该所有人拥有的专利集合的差集(集合相减)，获得有效的专利引用集合，获得中心度

# 9.增加所有人的集合体 公司，专利对象增加 引用集合属性

def add_company_to_group(com_name, group_name):
    global group_dict
    if group_name not in group_dict:
        new_com = GroupItem(group_name)
        group_dict[group_name] = new_com
    group_dict[group_name].add_company(com_name)


def add_patent_to_company(name_patent, company_set):
    global company_dict
    for company_name in company_set:
        if company_name not in company_dict:
            new_company = CompanyItem(company_name)
            company_dict[company_name] = new_company
        company_dict[company_name].addToPatentset(name_patent)


def add_patent_and_cite(name_patent, cite_ver):
    global patent_dict

    # 添加引用专利对象到字典中
    if name_patent not in patent_dict:
        new_patent = PatentItem(name_patent)
        patent_dict[name_patent] = new_patent
    patent_dict[name_patent].add_cite_set(cite_ver)


def add_patent_and_cited(name_patent, cite_ver):
    global patent_dict

    for cite_name in cite_ver:

        # 添加被引用专利对象到字典中 并增加其被引用属性
        if cite_name not in patent_dict:
            new_patent = PatentItem(cite_name)
            patent_dict[cite_name] = new_patent

        patent_dict[cite_name].add_cited_set(name_patent)


def set_unique_name(name_set, unique_name):
    global unique_name_dict
    for name_item in name_set:
        if name_item not in unique_name_dict:
            unique_name_dict[name_item] = unique_name


def get_unique_name(names):
    if isinstance(names, str):
        if names in unique_name_dict:
            return unique_name_dict[names]
        else:
            return names
    elif isinstance(names, set) or isinstance(names, list):
        result = set()
        for item in names:
            if item in unique_name_dict:
                result.add(unique_name_dict[names])
            else:
                result.add(item)
        return result
    else:
        return None


def get_centre_degree():
    global company_dict
    global patent_dict

    wb = Workbook()
    ws = wb.active
    row_index = 1
    # 遍历所有人
    for company_name, company_obj in company_dict.items():
        valid_set = set()
        # 遍历该所有人的全部专利
        for patent in company_obj.patentset:
            # 且该专利在专利字典中存在
            if patent in patent_dict:
                valid_set = valid_set | patent_dict[patent].cited_set
        valid_set = valid_set - company_obj.patentset
        # 获得中心度
        company_obj.centre_degrees = len(valid_set)
        ws.cell(row=row_index, column=1, value=company_name)
        ws.cell(row=row_index, column=2, value=company_obj.centredegrees)
        row_index += 1
        print("所有人: %-60s 中心度为: %-5s" % (company_name, company_obj.centredegrees))
    wb.save(output_path)


def cal_centre_degree():
    print('正在读取excel文件.................')
    wb = load_workbook(tar_excel_path)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    print('读取完成,开始数据处理.............')

    # 获得行数
    num_rows = ws.max_row

    # 获得标签行
    tag_row = [item.value for item in ws[1]]
    tag_row = [item.strip() for item in tag_row]

    # 获得关键数据列号
    patent_index = tag_row.index('GA') + 1  # 专利唯一名称
    company_index = tag_row.index('AE') + 1  # 所有人名称
    ref_index = tag_row.index('CP') + 1  # 引用专利
    subname_index = tag_row.index('PN') + 1  # 专利多个名称

    pattern_com = re.compile(r'\((.+?)\)')
    # 建立 所有人-唯一专利名字典 专利名-唯一专利名字典
    #      集团-所有人
    for i in range(2, num_rows):
        # 公司set
        company_str = str(ws.cell(row=i, column=company_index).value)
        company_set = set([item.strip() for item in company_str.split(';')])
        # 专利唯一名称
        patent_name = str(ws.cell(row=i, column=patent_index).value)
        # 专利子名称set
        subname_str = str(ws.cell(row=i, column=subname_index).value)
        subname_set = set([item.strip() for item in subname_str.split(';')])

        # 建立专利名-唯一专利名字典
        set_unique_name(subname_set, patent_name)

        # 建立所有人-唯一专利名字典
        add_patent_to_company(patent_name, company_set)

        # 建立集团-所有人字典
        for company in company_set:
            add_company_to_group(company, pattern_com.findall(company)[0])

        print('名称一致化处理已经完成%d/%d' % (i, num_rows))

    pattern = re.compile(r'([a-zA-Z0-9]+?-[A-Za-z]\d?)\s')
    # 建立专利-引用关系
    for j in range(2, num_rows):
        # 专利名
        patent_name = str(ws.cell(row=j, column=patent_index).value)
        # 引用专利名
        cite_set = set()
        cite_str = str(ws.cell(row=j, column=ref_index).value)
        if cite_str is None or cite_str.strip() == '':
            pass
        else:
            # 获取引用正则匹配列表
            cite_set = set(pattern.findall(cite_str)[1:])
            cite_set = get_unique_name(cite_set)

        add_patent_and_cite(patent_name, cite_set)
        add_patent_and_cited(patent_name, cite_set)

        print('引用关联已经完成%d/%d' % (i, num_rows))

    print('开始计算中心度:')
    # 获得中心度
    get_centre_degree()


if __name__ == '__main__':
    cal_centre_degree()
